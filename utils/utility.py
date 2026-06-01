# Databricks notebook source
# MAGIC %pip install openpyxl

# COMMAND ----------

dbutils.library.restartPython()

# COMMAND ----------

import os
import openpyxl
from openpyxl import load_workbook
import json
import tempfile

# COMMAND ----------

def read_json(spark, file_list):
    return spark.read.option("multiline", "true").json(file_list)

# COMMAND ----------

def read_xlsx(spark, file_list):


    path = file_list[0] if isinstance(file_list, list) else file_list
    local_path = path.replace("file:", "")

    wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows)]

    tmp = local_path + ".jsonl"
    with open(tmp, "w") as f:
        for row in rows:
            record = dict(zip(headers, [str(v) if v is not None else "" for v in row]))
            f.write(json.dumps(record) + "\n")

    wb.close()
    return spark.read.json(tmp) 

# COMMAND ----------

def read_csv(spark, file_list, header=True):
    return spark.read.option("header", header).csv(file_list)

# COMMAND ----------

# DBTITLE 1,m
def xlsx_to_jsonl(xlsx_path, jsonl_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows)]
    with open(jsonl_path, "w") as f:
        for row in rows:
            record = dict(zip(headers, [str(v) if v is not None else "" for v in row]))
            f.write(json.dumps(record) + "\n")
    wb.close()
    return jsonl_path

# COMMAND ----------

def write_table(spark, df, table_name, mode="overwrite"):
    df.write.mode(mode).option("mergeSchema", "true").saveAsTable(table_name)